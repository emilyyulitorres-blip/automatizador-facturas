from datetime import datetime
from io import BytesIO
from pathlib import Path
import os
import re
import unicodedata

import fitz
import pandas as pd
import pytesseract
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from PIL import Image, ImageEnhance, ImageFilter, ImageOps
from supabase import Client, create_client


st.set_page_config(
    page_title="Automatizador de Facturas",
    page_icon="📄",
    layout="wide",
)

BASE_DIR = Path(__file__).resolve().parent
RUC_SOLARTEAM = "1793069479001"

LOGO_CANDIDATES = [
    BASE_DIR / "logo_corregido.png",
]

if os.name == "nt":
    ruta_tesseract = Path(r"C:\Program Files\Tesseract-OCR\tesseract.exe")
    if ruta_tesseract.exists():
        pytesseract.pytesseract.tesseract_cmd = str(ruta_tesseract)


def preparar_imagen(imagen: Image.Image) -> Image.Image:
    imagen = ImageOps.exif_transpose(imagen).convert("L")
    imagen = ImageOps.autocontrast(imagen)
    imagen = ImageEnhance.Contrast(imagen).enhance(1.8)
    return imagen.filter(ImageFilter.SHARPEN)


def _puntaje_ocr(texto: str) -> int:
    texto = (texto or "").upper()
    palabras = ["FACTURA", "RUC", "TOTAL", "IVA", "FECHA", "SUBTOTAL"]
    return sum(texto.count(palabra) for palabra in palabras) + len(texto) // 200


def aplicar_ocr(imagen: Image.Image) -> str:
    preparada = preparar_imagen(imagen)
    resultados = []

    for angulo in (0, 90, 270, 180):
        rotada = preparada.rotate(angulo, expand=True)
        texto = pytesseract.image_to_string(
            rotada,
            lang="spa",
            config="--oem 3 --psm 6",
        )
        resultados.append((_puntaje_ocr(texto), texto))

    return max(resultados, key=lambda item: item[0])[1]


def leer_imagen(archivo) -> tuple[str, list[tuple]]:
    imagen = Image.open(BytesIO(archivo.getvalue()))
    return aplicar_ocr(imagen), []


def leer_pdf(archivo) -> tuple[str, list[tuple]]:
    texto_total = ""
    bloques_totales: list[tuple] = []
    datos = archivo.getvalue()

    with fitz.open(stream=datos, filetype="pdf") as documento:
        for pagina in documento:
            texto_pagina = pagina.get_text("text")
            bloques_totales.extend(pagina.get_text("blocks"))
            texto_total += texto_pagina + "\n"

            if len(texto_pagina.strip()) < 40:
                pixmap = pagina.get_pixmap(
                    matrix=fitz.Matrix(2.4, 2.4),
                    alpha=False,
                )
                imagen = Image.frombytes(
                    "RGB",
                    [pixmap.width, pixmap.height],
                    pixmap.samples,
                )
                texto_total += aplicar_ocr(imagen) + "\n"

    return texto_total, bloques_totales


def vista_previa(archivo) -> Image.Image:
    nombre = archivo.name.lower()

    if nombre.endswith((".jpg", ".jpeg", ".png")):
        return Image.open(BytesIO(archivo.getvalue()))

    with fitz.open(stream=archivo.getvalue(), filetype="pdf") as documento:
        pagina = documento[0]
        pixmap = pagina.get_pixmap(
            matrix=fitz.Matrix(1.4, 1.4),
            alpha=False,
        )
        return Image.frombytes(
            "RGB",
            [pixmap.width, pixmap.height],
            pixmap.samples,
        )

PATRON_FACTURA = re.compile(r"\b\d{3}-\d{3}-\d{9}\b")
PATRON_RUC = re.compile(r"\b\d{13}\b")
PATRON_FECHA = re.compile(
    r"\b(?:\d{2}[/-]\d{2}[/-]\d{4}|\d{4}[/-]\d{2}[/-]\d{2})\b"
)
PATRON_MONTO = re.compile(
    r"(?<!\d)(\d{1,3}(?:[.,]\d{3})*[.,]\d{2}|\d+[.,]\d{2})(?!\d)"
)


def normalizar(texto: str) -> str:
    texto = unicodedata.normalize("NFKD", texto or "")
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = texto.upper().replace("\xa0", " ")
    return re.sub(r"\s+", " ", texto).strip()


def lineas_limpias(texto: str) -> list[str]:
    return [
        re.sub(r"\s+", " ", linea).strip()
        for linea in texto.splitlines()
        if linea.strip()
    ]


def monto_normalizado(valor: str) -> str:
    if not valor:
        return ""

    valor = valor.replace("$", "").replace(" ", "")

    if "," in valor and "." in valor:
        if valor.rfind(",") > valor.rfind("."):
            valor = valor.replace(".", "").replace(",", ".")
        else:
            valor = valor.replace(",", "")
    else:
        valor = valor.replace(",", ".")

    try:
        return f"{float(valor):.2f}"
    except ValueError:
        return ""


def detectar_tipo(texto: str) -> str:
    n = normalizar(texto)
    compacto = re.sub(r"[^A-Z]", "", n)

    if "PROFORMA" in n or "PROFORMA" in compacto:
        return "NO VÁLIDA - PROFORMA"
    if "COTIZACION" in n:
        return "NO VÁLIDA - COTIZACIÓN"
    if "NOTA DE CREDITO" in n:
        return "NO VÁLIDA - NOTA DE CRÉDITO"
    if "NOTA DE VENTA" in n:
        return "REVISAR - NOTA DE VENTA"
    if "FACTURA" in n or "FACTURA" in compacto:
        if "NUMERO DE AUTORIZACION" in n or "CLAVE DE ACCESO" in n:
            return "FACTURA ELECTRÓNICA"
        return "FACTURA"

    return "REVISAR - NO IDENTIFICADO"


def extraer_factura(texto: str) -> str:
    coincidencias = PATRON_FACTURA.findall(texto)
    return coincidencias[0] if coincidencias else ""


def extraer_fecha(texto: str) -> str:
    lineas = lineas_limpias(texto)

    for indice, linea in enumerate(lineas):
        n = normalizar(linea)
        if (
            "FECHA EMISION" in n
            or "FECHA DE EMISION" in n
            or n == "FECHA"
            or n.startswith("FECHA ")
        ):
            fecha = PATRON_FECHA.search(linea)
            if fecha:
                return fecha.group(0)

            for siguiente in range(indice + 1, min(indice + 4, len(lineas))):
                fecha = PATRON_FECHA.search(lineas[siguiente])
                if fecha:
                    return fecha.group(0)

    fechas = PATRON_FECHA.findall(texto)
    return fechas[-1] if fechas else ""


def extraer_ruc_emisor(texto: str) -> str:
    rucs = PATRON_RUC.findall(texto)

    for ruc in rucs:
        if ruc != RUC_SOLARTEAM:
            return ruc

    return rucs[0] if rucs else ""


def extraer_cliente(texto: str) -> str:
    n = normalizar(texto)
    if "SOLARTEAM SAS" in n:
        return "SOLARTEAM SAS"
    if "SOLAR TEAM S.A.S" in n:
        return "SOLAR TEAM S.A.S"
    return ""


def _es_proveedor_valido(linea: str) -> bool:
    n = normalizar(linea)
    excluir = [
        "RUC", "FACTURA", "NUMERO DE AUTORIZACION", "CLAVE DE ACCESO",
        "FECHA", "AMBIENTE", "EMISION", "PRODUCCION", "NORMAL",
        "DIRECCION", "OBLIGADO", "CONTRIBUYENTE", "RAZON SOCIAL",
        "IDENTIFICACION", "SOLARTEAM", "SOLAR TEAM", "SUBTOTAL",
        "TOTAL", "CODIGO", "DESCRIPCION", "CANTIDAD", "PRECIO",
    ]

    return (
        len(linea.strip()) >= 7
        and not any(palabra in n for palabra in excluir)
        and not PATRON_RUC.search(linea)
        and not PATRON_FACTURA.search(linea)
        and not PATRON_FECHA.search(linea)
        and bool(re.search(r"[A-ZÁÉÍÓÚÑ]{3,}", linea.upper()))
    )


def extraer_proveedor(texto: str, ruc_emisor: str) -> str:
    lineas = lineas_limpias(texto)

    for indice, linea in enumerate(lineas):
        if ruc_emisor and ruc_emisor in linea:
            for siguiente in range(indice + 1, min(indice + 10, len(lineas))):
                candidato = lineas[siguiente].strip()
                if _es_proveedor_valido(candidato):
                    return candidato

    for linea in lineas[:45]:
        if _es_proveedor_valido(linea):
            return linea.strip()

    return ""


def _monto_desde_bloques(
    bloques: list[tuple],
    etiquetas: list[str],
    excluir: list[str] | None = None,
) -> str:
    excluir = excluir or []

    for etiqueta in etiquetas:
        etiqueta_n = normalizar(etiqueta)

        for bloque in bloques:
            texto_bloque = bloque[4] if len(bloque) > 4 else ""
            bloque_n = normalizar(texto_bloque)

            if etiqueta_n not in bloque_n:
                continue
            if any(normalizar(p) in bloque_n for p in excluir):
                continue

            montos = PATRON_MONTO.findall(texto_bloque)
            if montos:
                return monto_normalizado(montos[0])

    return ""


def _monto_en_texto(texto: str, patrones: list[str]) -> str:
    n = normalizar(texto)

    for patron in patrones:
        coincidencia = re.search(patron, n, re.IGNORECASE)
        if coincidencia:
            return monto_normalizado(coincidencia.group(1))

    return ""


def extraer_subtotal(texto: str, bloques: list[tuple]) -> str:
    valor = _monto_desde_bloques(
        bloques,
        ["SUBTOTAL SIN IMPUESTOS", "SUBTOTAL 15%", "SUBTOTAL 12%",
         "BASE IMPONIBLE", "BASE GRAVADA"],
        excluir=["NO OBJETO", "EXENTO", "DESCUENTO"],
    )
    return valor or _monto_en_texto(
        texto,
        [
            r"SUBTOTAL\s+SIN\s+IMPUESTOS\s*[:$]?\s*([0-9.,]+\d{2})",
            r"SUBTOTAL\s+15\s*%\s*[:$]?\s*([0-9.,]+\d{2})",
            r"SUBTOTAL\s+12\s*%\s*[:$]?\s*([0-9.,]+\d{2})",
            r"BASE\s+IMPONIBLE\s*[:$]?\s*([0-9.,]+\d{2})",
        ],
    )


def extraer_iva(texto: str, bloques: list[tuple]) -> str:
    valor = _monto_desde_bloques(
        bloques,
        ["IVA 15%", "IVA 12%", "TOTAL IVA", "IMPUESTO IVA"],
        excluir=["SUBTOTAL", "NO OBJETO", "EXENTO", "INCLUYE IVA"],
    )
    return valor or _monto_en_texto(
        texto,
        [
            r"\bIVA\s+15\s*%\s*[:$]?\s*([0-9.,]+\d{2})",
            r"\bIVA\s+12\s*%\s*[:$]?\s*([0-9.,]+\d{2})",
            r"TOTAL\s+IVA\s*[:$]?\s*([0-9.,]+\d{2})",
        ],
    )


def extraer_total(texto: str, bloques: list[tuple]) -> str:
    valor = _monto_desde_bloques(
        bloques,
        ["VALOR TOTAL", "TOTAL A PAGAR", "IMPORTE TOTAL",
         "MONTO TOTAL", "TOTAL FACTURA"],
        excluir=["SIN SUBSIDIO", "DESCUENTO", "SUBTOTAL", "AHORRO"],
    )
    return valor or _monto_en_texto(
        texto,
        [
            r"VALOR\s+TOTAL(?!\s+SIN\s+SUBSIDIO)\s*[:$]?\s*([0-9.,]+\d{2})",
            r"TOTAL\s+A\s+PAGAR\s*[:$]?\s*([0-9.,]+\d{2})",
            r"IMPORTE\s+TOTAL\s*[:$]?\s*([0-9.,]+\d{2})",
            r"MONTO\s+TOTAL\s*[:$]?\s*([0-9.,]+\d{2})",
        ],
    )


def validar(datos: dict) -> tuple[str, str]:
    if datos["Tipo documento"].startswith("NO VÁLIDA"):
        return "REVISAR", "Documento no registrable como factura"

    obligatorios = [
        "Fecha", "Factura", "Proveedor", "RUC Emisor",
        "Subtotal", "IVA", "Total",
    ]

    faltantes = [
        campo for campo in obligatorios
        if not str(datos.get(campo, "")).strip()
    ]
    if faltantes:
        return "REVISAR", "Faltan: " + ", ".join(faltantes)

    try:
        subtotal = float(datos["Subtotal"])
        iva = float(datos["IVA"])
        total = float(datos["Total"])
        if abs((subtotal + iva) - total) > 0.10:
            return "REVISAR", "Subtotal + IVA no coincide con el total"
    except ValueError:
        return "REVISAR", "Los valores monetarios requieren revisión"

    return "OK", ""


def calcular_confianza(datos: dict) -> int:
    campos = ["Fecha", "Factura", "Proveedor", "RUC Emisor", "Subtotal", "IVA", "Total"]
    completos = sum(bool(str(datos.get(campo, "")).strip()) for campo in campos)
    confianza = int((completos / len(campos)) * 100)

    if datos["Tipo documento"].startswith("NO VÁLIDA"):
        confianza = min(confianza, 40)

    return confianza


def extraer_datos(texto: str, bloques: list[tuple], nombre_archivo: str) -> dict:
    tipo = detectar_tipo(texto)
    ruc_emisor = extraer_ruc_emisor(texto)

    datos = {
        "Estado": "",
        "Confianza": 0,
        "Tipo documento": tipo,
        "Archivo": nombre_archivo,
        "Fecha": extraer_fecha(texto),
        "Factura": extraer_factura(texto),
        "Proveedor": extraer_proveedor(texto, ruc_emisor),
        "RUC Emisor": ruc_emisor,
        "Cliente": extraer_cliente(texto),
        "RUC Cliente": RUC_SOLARTEAM if RUC_SOLARTEAM in texto else "",
        "Subtotal": extraer_subtotal(texto, bloques),
        "IVA": extraer_iva(texto, bloques),
        "Total": extraer_total(texto, bloques),
        "Proyecto": "",
        "Consumo": "",
        "Categoría": "",
        "Observación": "",
    }

    estado, observacion = validar(datos)
    datos["Estado"] = estado
    datos["Observación"] = observacion
    datos["Confianza"] = calcular_confianza(datos)

    return datos


def marcar_duplicados(df):
    if df.empty or "Factura" not in df.columns:
        return df

    duplicados = (
        df["Factura"].astype(str).str.strip().ne("")
        & df["Factura"].astype(str).duplicated(keep=False)
    )

    for indice in df.index[duplicados]:
        df.at[indice, "Estado"] = "REVISAR"
        observacion = str(df.at[indice, "Observación"] or "").strip()
        mensaje = "Factura duplicada en la carga"
        df.at[indice, "Observación"] = f"{observacion}; {mensaje}".strip("; ")
        df.at[indice, "Confianza"] = min(int(df.at[indice, "Confianza"]), 70)

    return df

@st.cache_resource
def obtener_supabase() -> Client | None:
    try:
        return create_client(
            st.secrets["SUPABASE_URL"],
            st.secrets["SUPABASE_KEY"],
        )
    except Exception:
        return None


def cargar_memoria(cliente: Client | None) -> dict:
    if cliente is None:
        return {}

    try:
        respuesta = (
            cliente.table("correcciones")
            .select("id,ruc_emisor,campo,valor_corregido")
            .order("id", desc=True)
            .limit(1000)
            .execute()
        )

        memoria = {}
        for fila in respuesta.data or []:
            clave = (
                str(fila.get("ruc_emisor") or "").strip(),
                str(fila.get("campo") or "").strip(),
            )
            if clave not in memoria:
                memoria[clave] = str(fila.get("valor_corregido") or "").strip()

        return memoria
    except Exception:
        return {}


def aplicar_memoria(datos: dict, memoria: dict) -> dict:
    ruc = str(datos.get("RUC Emisor") or "").strip()

    for campo in ("Proveedor", "Categoría", "Consumo"):
        valor = memoria.get((ruc, campo))
        if valor:
            datos[campo] = valor

    return datos


def obtener_o_crear_proyecto(
    cliente: Client,
    nombre_proyecto: str,
) -> int | None:
    nombre = re.sub(r"\s+", " ", str(nombre_proyecto or "").strip())
    if not nombre:
        return None

    try:
        respuesta = (
            cliente.table("proyectos")
            .select("id,nombre")
            .ilike("nombre", nombre)
            .limit(1)
            .execute()
        )

        if respuesta.data:
            return int(respuesta.data[0]["id"])

        creado = (
            cliente.table("proyectos")
            .insert({"nombre": nombre, "estado": "Activo"})
            .execute()
        )

        if creado.data:
            return int(creado.data[0]["id"])
    except Exception:
        return None

    return None


def _fecha_bd(valor) -> str | None:
    texto = str(valor or "").strip()
    if not texto:
        return None

    for formato in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(texto, formato).date().isoformat()
        except ValueError:
            continue

    return None


def _numero(valor):
    texto = str(valor or "").strip()
    if not texto:
        return None
    try:
        return float(texto.replace(",", "."))
    except ValueError:
        return None


def guardar_facturas(
    cliente: Client,
    df_original: pd.DataFrame,
    df_editado: pd.DataFrame,
) -> tuple[int, int, list[str]]:
    guardadas = 0
    duplicadas = 0
    errores: list[str] = []

    for indice, fila in df_editado.iterrows():
        proyecto_nombre = str(fila.get("Proyecto") or "").strip()
        proyecto_id = None

        if proyecto_nombre:
            proyecto_id = obtener_o_crear_proyecto(cliente, proyecto_nombre)
            if not proyecto_id:
                errores.append(
                    f"{fila.get('Archivo', 'Documento')}: "
                    "no se pudo crear o encontrar el proyecto."
                )
                continue

        registro = {
            "fecha": _fecha_bd(fila.get("Fecha")),
            "numero_factura": str(fila.get("Factura") or "").strip(),
            "proveedor": str(fila.get("Proveedor") or "").strip(),
            "ruc_emisor": str(fila.get("RUC Emisor") or "").strip(),
            "cliente": str(fila.get("Cliente") or "").strip(),
            "ruc_cliente": str(fila.get("RUC Cliente") or "").strip(),
            "subtotal": _numero(fila.get("Subtotal")),
            "iva": _numero(fila.get("IVA")),
            "total": _numero(fila.get("Total")),
            "proyecto_id": proyecto_id,
            "consumo": str(fila.get("Consumo") or "").strip(),
            "categoria": str(fila.get("Categoría") or "").strip(),
            "archivo": str(fila.get("Archivo") or "").strip(),
            "estado": str(fila.get("Estado") or "").strip(),
            "observacion": str(fila.get("Observación") or "").strip(),
            "confianza": int(fila.get("Confianza") or 0),
        }

        if not registro["numero_factura"] or not registro["ruc_emisor"]:
            errores.append(
                f"{registro['archivo']}: falta factura o RUC emisor."
            )
            continue

        try:
            cliente.table("facturas").insert(registro).execute()
            guardadas += 1
        except Exception as error:
            mensaje = str(error).lower()
            if "duplicate" in mensaje or "unique" in mensaje or "23505" in mensaje:
                duplicadas += 1
            else:
                errores.append(f"{registro['archivo']}: {error}")
                continue

        if indice not in df_original.index:
            continue

        original = df_original.loc[indice]
        ruc = registro["ruc_emisor"]

        for campo in ("Proveedor", "Categoría", "Consumo"):
            detectado = str(original.get(campo) or "").strip()
            corregido = str(fila.get(campo) or "").strip()

            if corregido and corregido != detectado:
                try:
                    cliente.table("correcciones").insert({
                        "ruc_emisor": ruc,
                        "campo": campo,
                        "valor_detectado": detectado,
                        "valor_corregido": corregido,
                    }).execute()
                except Exception:
                    pass

    return guardadas, duplicadas, errores

def crear_excel(df: pd.DataFrame) -> bytes:
    salida = BytesIO()

    with pd.ExcelWriter(salida, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Facturas")

    salida.seek(0)
    libro = load_workbook(salida)
    hoja = libro["Facturas"]

    for celda in hoja[1]:
        celda.fill = PatternFill(fill_type="solid", fgColor="172B4D")
        celda.font = Font(color="FFFFFF", bold=True)
        celda.alignment = Alignment(horizontal="center", vertical="center")

    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = hoja.dimensions
    hoja.row_dimensions[1].height = 24

    for columna in hoja.columns:
        ancho = min(
            max(len(str(celda.value or "")) for celda in columna) + 2,
            45,
        )
        hoja.column_dimensions[columna[0].column_letter].width = ancho

    if hoja.max_row >= 2:
        referencia = f"A1:{hoja.cell(hoja.max_row, hoja.max_column).coordinate}"
        tabla = Table(displayName="TablaFacturas", ref=referencia)
        tabla.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        hoja.add_table(tabla)

    final = BytesIO()
    libro.save(final)
    return final.getvalue()

def aplicar_estilos() -> None:
    st.markdown(
        """
        <style>
        .block-container {
            max-width: 1550px;
            padding-top: 1.3rem;
            padding-bottom: 2rem;
        }
        .section-title {
            font-size: 1.5rem;
            font-weight: 750;
            color: #172B4D;
            margin-top: 1.5rem;
            margin-bottom: 0.8rem;
        }
        .footer {
            margin-top: 2rem;
            padding: 18px 22px;
            border-radius: 14px;
            background: #F8FAFC;
            color: #64748B;
            font-size: 0.9rem;
        }
        [data-testid="stFileUploader"] {
            border: 1px dashed #94A3B8;
            border-radius: 14px;
            padding: 0.4rem;
            background: #F8FAFC;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def mostrar_encabezado() -> None:
    logo = next((ruta for ruta in LOGO_CANDIDATES if ruta.exists()), None)

    col_logo, col_texto = st.columns([1.4, 5.6], vertical_alignment="center")

    with col_logo:
        if logo:
            # Streamlit nativo: evita que el HTML se muestre como texto.
            st.image(str(logo), width=180)
        else:
            st.caption("Logo no encontrado")

    with col_texto:
        st.title("Automatizador de Facturas")
        st.caption(
            "Carga facturas PDF o imágenes, revisa la información extraída "
            "y genera un Excel listo para descargar."
        )

    st.divider()


def titulo_seccion(texto: str) -> None:
    st.markdown(
        f'<div class="section-title">{texto}</div>',
        unsafe_allow_html=True,
    )


def pie_pagina() -> None:
    st.markdown(
        """
        <div class="footer">
            <b>Solar Team</b><br>
            Automatización de información de facturas · Versión 4.0
        </div>
        """,
        unsafe_allow_html=True,
    )

aplicar_estilos()
mostrar_encabezado()

supabase = obtener_supabase()
memoria = cargar_memoria(supabase)

if supabase is None:
    st.warning(
        "La extracción y el Excel funcionan, pero no se pudo conectar "
        "con Supabase. Revisa los Secrets de Streamlit."
    )

titulo_seccion("📂 Sube tus facturas")

archivos = st.file_uploader(
    "Arrastra aquí tus archivos o selecciónalos",
    type=["pdf", "jpg", "jpeg", "png"],
    accept_multiple_files=True,
    help="Formatos permitidos: PDF, JPG, JPEG y PNG",
)

if archivos:
    resultados = []
    barra = st.progress(0, text="Preparando documentos...")

    for indice, archivo in enumerate(archivos, start=1):
        try:
            if archivo.name.lower().endswith(".pdf"):
                texto, bloques = leer_pdf(archivo)
            else:
                texto, bloques = leer_imagen(archivo)

            datos = extraer_datos(texto, bloques, archivo.name)
            resultados.append(aplicar_memoria(datos, memoria))

        except Exception as error:
            resultados.append({
                "Estado": "ERROR",
                "Confianza": 0,
                "Tipo documento": "ERROR",
                "Archivo": archivo.name,
                "Fecha": "",
                "Factura": "",
                "Proveedor": "",
                "RUC Emisor": "",
                "Cliente": "",
                "RUC Cliente": "",
                "Subtotal": "",
                "IVA": "",
                "Total": "",
                "Proyecto": "",
                "Consumo": "",
                "Categoría": "",
                "Observación": str(error),
            })

        barra.progress(
            indice / len(archivos),
            text=f"Procesando {indice} de {len(archivos)} documentos...",
        )

    barra.empty()

    df = marcar_duplicados(pd.DataFrame(resultados))
    df_original = df.copy(deep=True)

    titulo_seccion("📊 Resumen del procesamiento")
    c1, c2, c3, c4 = st.columns(4)

    c1.metric("✅ Facturas OK", int((df["Estado"] == "OK").sum()))
    c2.metric("⚠️ Para revisar", int((df["Estado"] == "REVISAR").sum()))
    c3.metric("❌ Errores", int((df["Estado"] == "ERROR").sum()))

    total_detectado = pd.to_numeric(df["Total"], errors="coerce").fillna(0).sum()
    c4.metric("💰 Total detectado", f"${total_detectado:,.2f}")

    titulo_seccion("📋 Revisa y corrige antes de descargar")

    df_editado = st.data_editor(
        df,
        use_container_width=True,
        hide_index=True,
        num_rows="dynamic",
        height=min(640, 140 + len(df) * 42),
        disabled=["Archivo"],
        column_config={
            "Estado": st.column_config.SelectboxColumn(
                "Estado",
                options=["OK", "REVISAR", "ERROR"],
                width="small",
            ),
            "Confianza": st.column_config.ProgressColumn(
                "Confianza",
                min_value=0,
                max_value=100,
                format="%d%%",
                width="small",
            ),
            "Archivo": st.column_config.TextColumn(width="large"),
            "Factura": st.column_config.TextColumn(width="medium"),
            "Proveedor": st.column_config.TextColumn(width="large"),
            "Proyecto": st.column_config.TextColumn(
                "Proyecto",
                width="medium",
                help=(
                    "Escribe el proyecto o déjalo vacío. "
                    "Los proyectos nuevos se crean al guardar."
                ),
            ),
            "Observación": st.column_config.TextColumn(width="large"),
        },
    )

    nombre_excel = "FACTURAS_" + datetime.now().strftime("%Y-%m-%d_%H-%M") + ".xlsx"

    col_excel, col_guardar = st.columns(2)

    with col_excel:
        st.download_button(
            "📥 Descargar Excel",
            data=crear_excel(df_editado),
            file_name=nombre_excel,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    with col_guardar:
        guardar = st.button(
            "💾 Guardar en base de datos",
            type="primary",
            use_container_width=True,
            disabled=(supabase is None),
        )

    if guardar and supabase is not None:
        with st.spinner("Guardando facturas y correcciones..."):
            guardadas, duplicadas, errores = guardar_facturas(
                supabase,
                df_original,
                df_editado,
            )

        if guardadas:
            st.success(f"Se guardaron {guardadas} factura(s) correctamente.")
        if duplicadas:
            st.warning(f"{duplicadas} factura(s) ya existían y no se duplicaron.")
        if errores:
            st.error(
                "No se guardaron algunos registros:\n\n"
                + "\n".join(f"- {error}" for error in errores)
            )

    with st.expander("👁️ Vista previa de los documentos"):
        nombre_vista = st.selectbox(
            "Selecciona un archivo",
            [archivo.name for archivo in archivos],
        )
        seleccionado = next(
            archivo for archivo in archivos if archivo.name == nombre_vista
        )
        st.image(vista_previa(seleccionado), use_container_width=True)

else:
    st.info("Carga uno o varios archivos PDF, JPG, JPEG o PNG para comenzar.")

pie_pagina()
